#!/usr/bin/python
# -*- coding: UTF-8 -*-
import csv
import datetime
import os
import json
import string
import pandas as pd

from openpyxl import load_workbook
import logging
from io import StringIO

from db_tools import DBTools


class ExcelRunner:
    def __init__(self, task_config, database, logger, dict_tables, dict_id_map, dict_name_map) -> None:
        self.task_config = task_config
        self.database = database
        self.dict_name_map = dict_name_map
        self.dict_id_map = dict_id_map
        self.schema_name = None
        self.dbtools = None
        self.logger = logger
        self.dict_tables = dict_tables

        self.file_path = task_config['excel_file']
        self.base_column = task_config['base_column']
        self.fk_id_column = task_config['fk_id_column']
        self.ext_column = task_config['ext_column']
        self.default_column = task_config['default_column']
    def get_connection(self):
        return self.dbtools.get_connection()

    # 清空表数据
    def truncate_table(self, table_name):
        sql = f'TRUNCATE TABLE "{self.schema_name}"."{table_name}"'
        with self.get_connection() as conn:
            with conn.cursor() as cursor:
                try:
                    cursor.execute(sql)
                    conn.commit()
                except Exception as e:
                    self.logger.error(f"清空[{table_name}]数据失败", exc_info=True)
    # 插入基础数据
    def insert_base_data(self, data):
        col_names = ""
        col_values = ""

        # 清空表数据
        self.truncate_table(self.table_name)

        # 拼接基础字段
        for column_info in self.base_column:
            col_name = column_info['name']
            col_names = col_names + f'"{col_name}",'
            col_values = col_values + f'%({col_name})s,'               
        
        # 拼接默认字段
        for column_info in self.default_column:
            col_name = column_info['name']
            col_names = col_names + f'"{col_name}",'
            col_values = col_values + f'%({col_name})s,'

        # 拼接外键字段
        for column_info in self.fk_id_column:
            col_name = column_info['name']
            col_names = col_names + f'"{col_name}",'
            col_values = col_values + f'%({col_name})s,'

        sql = f'''
        INSERT INTO "{self.schema_name}"."{self.table_name}" (
        {col_names[:-1]})
        VALUES( {col_values[:-1]}
        );

        '''

        self.insert_data(sql, data)


    # 插入扩展数据
    def insert_ext_data(self, data):

        # 清空表数据
        self.truncate_table(self.extend_table)
        
        col_names = ""
        col_values = ""
        for column_info in self.ext_column:
            col_name = column_info['name']
            col_names = col_names + f'"{col_name}",'
            col_values = col_values + f'%({col_name})s,'
        sql = f'''
        INSERT INTO "{self.schema_name}"."{self.extend_table}" ( id, 
        {col_names[:-1]} )
        VALUES( %(id)s,
        {col_values[:-1]} 
        );

        '''
        self.insert_data(sql, data)


    # 插入数据公共方法
    def insert_data(self, sql, data):
        print(sql)
        with self.get_connection() as conn:
            with conn.cursor() as cursor:
                try:
                    conn.autocommit = False
                    cursor.execute("BEGIN;")
                    cursor.executemany(sql, data)
                    conn.commit()
                except Exception as e:
                    conn.rollback()
                    self.logger.error("执行executemany失败", exc_info=True)
                finally:
                    # conn.autocommit = True
                    pass


    # 快速插入扩展数据
    def fast_insert_ext_data(self, data, task_config):
        extend_table = task_config['extend_table']
        ext_column = task_config['ext_column']
        
        col_names = []
        for column_info in ext_column:
            col_name = column_info['name']
            col_names.append(col_name)

        # 将字典列表转换为CSV格式的字符串
        csv_data = '' 
        for row in data:
            row_str = ''
            for col_name in col_names:
                col_value = row[col_name]
                if type(col_value) == str and any(char in string.punctuation for char in col_value):
                    row_str = row_str + f'"{col_value}",'
                else:
                    row_str = row_str + f'{col_value},'
            csv_data = csv_data + row_str[:-1] + '\n'

        # 连接到数据库
        with self.get_connection() as conn:
            with conn.cursor() as cur:
                try:
                    conn.autocommit = False
                    # 开启事务
                    cur.execute("BEGIN;")
                    # 创建一个内存中的CSV文件对象
                    csv_file = StringIO(csv_data)
                    with open('output.csv', 'w', newline='', encoding='utf-8') as file:
                        # 写入数据到文件，注意我们直接使用getvalue()获取StringIO中的字符串内容
                        file.write(csv_file.getvalue())
                    # 使用copy_from方法批量导入数据
                    cur.copy_from(csv_file, f'{extend_table}', sep=',', null='', columns=col_names)

                    # 提交事务
                    conn.commit()

                except Exception as e:
                    # 如果发生错误，则回滚事务
                    conn.rollback()
                    self.logger.error(f"An error occurred: {e}")


    # 执行任务
    def run(self, task_id):
        self.logger.info('正在执行：%s', self.task_config.get('title', task_id))
        db = self.task_config['db']
        self.dbtools = DBTools(self.logger, self.database[db])
        self.schema_name = self.dbtools.get_schema_name()

        workbook = load_workbook(self.file_path)
        sheet_names = workbook.sheetnames
        row_num = 0

        base_data = [] 
        ext_data = [] 

        for sheet_name in sheet_names:
            self.logger.info('正在处理sheet页：%s', sheet_name)
            sheet = workbook[sheet_name]
            base_column_max_index = max(self.base_column, key=lambda x: x.get('index', 0))['index']

            # 扩展列名称数组
            ext_col_names = []
            first_row_data = [cell.value for cell in sheet[1]]
            for ext_col_num in range(base_column_max_index + 1, len(first_row_data)):
                    ext_col_name = first_row_data[ext_col_num]
                    if ext_col_name:
                        ext_col_names.append(ext_col_name)
            

            # 遍历sheet页中的行
            for row in sheet.iter_rows(min_row=2, values_only=True):
                
                eqled_term_info = {}
                eqled_term_info["id"] = row_num

                # 设置基础字段
                for column_item in self.base_column:
                    name = column_item['name']
                    index = column_item['index']
                    eqled_term_info[name] = row[index]
                    # self.logger.info(col_name, "=>",row[col_list.index(col_name)] )

                # 设置默认字段
                for column_info in self.default_column:
                    col_name = column_info['name']
                    if column_info.__contains__('value'):
                        col_value = column_info['value']
                        if(col_value == 'sheet_name'):
                            eqled_term_info[col_name] = sheet_name
                            continue
                        eqled_term_info[col_name] = col_value
                
                # 设置外键编号字段
                for column_info in self.fk_id_column:
                    col_name = column_info['name']
                    dict_table = column_info['dict_table']
                    
                    name_field = column_info.get('name_field', None)
                    id_field = column_info.get('id_field', None)

                    if name_field is not None:
                        value_prefix = column_info.get('value_prefix', '')
                        try:
                            name_value = eqled_term_info[name_field]
                            fk_id_value = self.dict_name_map[dict_table][value_prefix + name_value]
                            eqled_term_info[col_name] = fk_id_value
                        except:
                            eqled_term_info[col_name] = ''
                            continue
                    elif id_field is not None:
                        try:
                            id_value = eqled_term_info[id_field]
                            fk_name_value = self.dict_id_map[dict_table][id_value]
                            eqled_term_info[col_name] = fk_name_value
                        except:
                            eqled_term_info[col_name] = ''
                            continue

                base_data.append(eqled_term_info) 

                ################################基础字典完，开始扩展字段初始化#################################
                ext_col_count = 0 
                for ext_col_num in range(base_column_max_index + 1, len(first_row_data)):
                    if ext_col_count < len(ext_col_names):
                        ext_col_value = {}   
                        for ext_col_info in self.ext_column:
                            col_name = ext_col_info['name']
                            col_value = ext_col_info['value']
                            if eqled_term_info.__contains__(col_value):
                                value = eqled_term_info[col_value]
                            elif col_value == 'name':
                                value = ext_col_names[ext_col_count]
                            elif col_value == 'value':
                                value = row[ext_col_num]
                            else:
                                value = col_value
                            ext_col_value[col_name] = value
                        ext_col_value['id'] = ext_col_count
                        ext_data.append(ext_col_value)
                        ext_col_count = ext_col_count + 1
                row_num = row_num + 1
        
        self.logger.info("正在插入基础数据...")
        self.insert_base_data(base_data)
        self.logger.info("正在插入扩展数据...")
        self.insert_ext_data(ext_data)